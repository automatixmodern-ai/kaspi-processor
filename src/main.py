"""
main.py — Kaspi receipt processor (GitHub Actions edition).

Flow:
  1. Connect to Gmail over IMAP using an App Password (from env / GitHub Secrets).
  2. Find UNSEEN emails whose subject contains "Kaspi" and that have a .zip.
  3. Unzip (folders preserved). Skip Excel files. OCR PDFs/images.
  4. Kaspi documents  -> "Kaspi" sheet  (Folder, Document, Date, Name, Amount, Receipt No)
     Non-Kaspi docs   -> "No Kaspi" sheet (Folder, Document)
  5. Validation sheet: documents processed vs rows written.
  6. Reply to the ORIGINAL SENDER with the .xlsx attached, then mark the email Seen.

No per-day OCR cap (Tesseract runs locally on the runner), no 6-minute limit,
no credit card. Runs free on a public GitHub repo via a scheduled workflow.
"""

import os
import re
import sys
import ssl
import zipfile
import tempfile
import imaplib
import smtplib
import email
from email.message import EmailMessage
from email.header import decode_header, make_header
from email.utils import parseaddr
from datetime import datetime

import requests
import openpyxl
from openpyxl.styles import Font

from parser import parse_document

# ---- Config (via environment / GitHub Secrets) -----------------------------
GMAIL_USER = os.environ["GMAIL_USER"]           # e.g. automatixmodern@gmail.com
GMAIL_APP_PASSWORD = os.environ["GMAIL_APP_PASSWORD"]   # 16-char app password
GCP_API_KEY = os.environ.get("GCP_API_KEY", "")        # also used to download Drive files
SUBJECT_KEYWORD = os.environ.get("SUBJECT_KEYWORD", "Kaspi")
CONF_THRESHOLD = float(os.environ.get("CONF_THRESHOLD", "0.5"))  # flag fields below this

EXCLUDE_EXT = {"xls", "xlsx", "xlsm", "xltx", "csv"}
IMAGE_EXT = {"jpg", "jpeg", "jfif", "png", "gif", "bmp", "webp", "tif", "tiff"}
PDF_EXT = {"pdf"}

IMAP_HOST, IMAP_PORT = "imap.gmail.com", 993
SMTP_HOST, SMTP_PORT = "smtp.gmail.com", 465


def _decode(s):
    try:
        return str(make_header(decode_header(s)))
    except Exception:
        return s or ""


def is_junk(name):
    return ("__MACOSX" in name) or re.search(r'(^|/)\.', name) or name.endswith("/")


# ---- Google Drive link handling -------------------------------------------
# Matches the common Drive URL shapes and captures the file ID:
#   https://drive.google.com/file/d/<ID>/view?usp=sharing
#   https://drive.google.com/open?id=<ID>
#   https://drive.google.com/uc?id=<ID>&export=download
#   https://docs.google.com/.../d/<ID>/...
DRIVE_ID_PATTERNS = [
    re.compile(r'drive\.google\.com/file/d/([A-Za-z0-9_-]{20,})'),
    re.compile(r'drive\.google\.com/open\?id=([A-Za-z0-9_-]{20,})'),
    re.compile(r'drive\.google\.com/uc\?[^"\s]*id=([A-Za-z0-9_-]{20,})'),
    re.compile(r'docs\.google\.com/[^"\s]*/d/([A-Za-z0-9_-]{20,})'),
]


def _get_email_text(msg):
    """Concatenate all text/plain and text/html body parts into one string."""
    chunks = []
    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue
        if part.get_filename():           # skip attachments
            continue
        ctype = part.get_content_type()
        if ctype in ("text/plain", "text/html"):
            payload = part.get_payload(decode=True)
            if payload:
                try:
                    chunks.append(payload.decode(part.get_content_charset() or "utf-8",
                                                 errors="replace"))
                except Exception:
                    chunks.append(payload.decode("utf-8", errors="replace"))
    return "\n".join(chunks)


def find_drive_file_ids(text):
    """Return a de-duplicated list of Drive file IDs found in the email body."""
    ids = []
    for pat in DRIVE_ID_PATTERNS:
        for m in pat.findall(text or ""):
            if m not in ids:
                ids.append(m)
    return ids


def download_drive_file(file_id):
    """
    Download a Drive file shared as 'Anyone with the link' using the API key.
    Returns (filename, bytes). Raises RuntimeError with a clear message on failure.
    """
    if not GCP_API_KEY:
        raise RuntimeError("No GCP_API_KEY set for Drive download.")

    # 1) get the real filename (nice-to-have; falls back to the ID)
    fname = f"{file_id}.zip"
    try:
        meta = requests.get(
            f"https://www.googleapis.com/drive/v3/files/{file_id}",
            params={"key": GCP_API_KEY, "fields": "name,size"}, timeout=60)
        if meta.status_code == 200:
            fname = meta.json().get("name", fname)
    except Exception:
        pass

    # 2) download the bytes
    r = requests.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}",
        params={"key": GCP_API_KEY, "alt": "media"}, timeout=600, stream=True)
    if r.status_code == 403:
        raise RuntimeError("DRIVE_FORBIDDEN: the file is not shared as "
                           "'Anyone with the link' (or Drive API/key not permitted).")
    if r.status_code == 404:
        raise RuntimeError("DRIVE_NOT_FOUND: the Drive link/file ID is invalid.")
    r.raise_for_status()
    return fname, r.content


# ---- Excel builder ---------------------------------------------------------
def build_excel(kaspi_rows, nokaspi_rows, total_docs, excluded, out_path):
    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    s1 = wb.active
    s1.title = "Kaspi"
    s1.append(["Folder", "Document", "Date", "Name", "Amount", "Receipt Number"])
    for c in s1[1]:
        c.font = bold
    for row in kaspi_rows:
        s1.append(row)
    # keep long receipt numbers + dates as text (no scientific notation / auto-dates)
    for r in range(2, s1.max_row + 1):
        s1.cell(r, 3).number_format = "@"
        s1.cell(r, 6).number_format = "@"

    s2 = wb.create_sheet("No Kaspi")
    s2.append(["Folder", "Document"])
    for c in s2[1]:
        c.font = bold
    for row in nokaspi_rows:
        s2.append(row)

    rows_written = len(kaspi_rows) + len(nokaspi_rows)
    s3 = wb.create_sheet("Validation")
    for row in [
        ["Documents in folders (Excel excluded)", total_docs],
        ["Excel files excluded", excluded],
        ['Rows in "Kaspi" tab', len(kaspi_rows)],
        ['Rows in "No Kaspi" tab', len(nokaspi_rows)],
        ["Validation result", "PASSED" if total_docs == rows_written else "MISMATCH — CHECK"],
    ]:
        s3.append(row)
    for c in s3["A"]:
        c.font = bold

    wb.save(out_path)
    return rows_written


# ---- Process one email -----------------------------------------------------
ZIP_CONTENT_TYPES = {
    "application/zip", "application/x-zip-compressed", "application/x-zip",
    "application/octet-stream",   # some clients send zips as this
    "multipart/x-zip",
}


def _looks_like_zip(filename, content_type, payload):
    """True if this attachment is a zip, by name, MIME type, or magic bytes."""
    name = (filename or "").lower()
    if name.endswith(".zip"):
        return True
    if (content_type or "").lower() in ZIP_CONTENT_TYPES:
        # confirm with the zip magic number to avoid false positives on octet-stream
        if payload and payload[:2] == b"PK":
            return True
    # last resort: any attachment whose bytes start with the zip signature
    if payload and payload[:4] in (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08"):
        return True
    return False


def process_message(raw_bytes):
    """Returns a result_dict, or None if the message has no zip to act on."""
    msg = email.message_from_bytes(raw_bytes)
    sender_name, sender_addr = parseaddr(msg.get("From", ""))
    subject = _decode(msg.get("Subject", ""))

    # collect ALL zip attachments (there may be several — one per folder)
    zips = []
    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue
        fn = part.get_filename()
        if fn is None:
            continue
        fn = _decode(fn)
        payload = part.get_payload(decode=True)
        ctype = part.get_content_type()
        if payload and _looks_like_zip(fn, ctype, payload):
            zips.append((fn, payload))
            print(f"    zip attachment: {fn!r} ({ctype}, {len(payload)} bytes)")
        elif fn:
            print(f"    (skipped non-zip attachment: {fn!r} [{ctype}])")

    print(f"  Found {len(zips)} zip attachment(s) on this email.")

    # Also handle Google Drive links in the body (for files too big to attach).
    drive_errors = []
    body_text = _get_email_text(msg)
    file_ids = find_drive_file_ids(body_text)
    if file_ids:
        print(f"  Found {len(file_ids)} Google Drive link(s) in the body.")
    for fid in file_ids:
        try:
            dname, dbytes = download_drive_file(fid)
            # only keep it if it's actually a zip
            if dbytes[:2] == b"PK" or dname.lower().endswith(".zip"):
                zips.append((dname, dbytes))
                print(f"    downloaded from Drive: {dname!r} ({len(dbytes)} bytes)")
            else:
                print(f"    (Drive file {dname!r} is not a zip — skipped)")
        except RuntimeError as e:
            drive_errors.append(str(e))
            print(f"    [Drive download failed] {e}", file=sys.stderr)

    if not zips:
        # Nothing to process. If a Drive link was present but failed, tell the sender.
        if drive_errors:
            return {"drive_error_only": True, "to_addr": sender_addr,
                    "subject": subject, "errors": drive_errors, "msg_id": msg.get("Message-ID", "")}
        return None

    kaspi_rows, nokaspi_rows, flagged = [], [], []
    total_docs, excluded = 0, 0

    with tempfile.TemporaryDirectory() as tmp:
        for idx, (zname, zbytes) in enumerate(zips):
            zpath = os.path.join(tmp, f"in_{idx}.zip")
            with open(zpath, "wb") as f:
                f.write(zbytes)
            try:
                zf = zipfile.ZipFile(zpath)
            except zipfile.BadZipFile:
                continue

            for info in zf.infolist():
                name = info.filename
                if info.is_dir() or is_junk(name):
                    continue
                parts = name.split("/")
                doc = parts[-1]
                # Folder = internal path if present, else the zip file's own name
                # (sender sends one zip per folder, named after that folder).
                if len(parts) > 1:
                    folder = "/".join(parts[:-1])
                else:
                    folder = os.path.splitext(os.path.basename(zname))[0]
                ext = doc.rsplit(".", 1)[-1].lower() if "." in doc else ""

                if ext in EXCLUDE_EXT:
                    excluded += 1
                    continue
                if ext not in IMAGE_EXT and ext not in PDF_EXT:
                    # unknown format — still count it, send to No-Kaspi for visibility
                    total_docs += 1
                    nokaspi_rows.append([folder, doc])
                    continue

                total_docs += 1
                # extract this one file to disk for OCR
                fpath = os.path.join(tmp, "doc." + ext)
                with open(fpath, "wb") as out:
                    out.write(zf.read(info))

                try:
                    fields, conf, has_kaspi = parse_document(fpath, is_pdf=(ext in PDF_EXT))
                except Exception as e:
                    kaspi_rows.append([folder, doc, "", f"ERROR: {e}", "", ""])
                    flagged.append(f"{folder}/{doc}  (could not read: {e})")
                    continue

                if has_kaspi:
                    kaspi_rows.append([folder, doc, fields["date"], fields["name"],
                                       fields["amount"], fields["receiptNo"]])
                    missing = [k for k in ("date", "name", "amount", "receiptNo")
                               if not fields[k] or conf[k] < CONF_THRESHOLD]
                    if missing:
                        flagged.append(f"{folder}/{doc}  (empty/uncertain: {', '.join(missing)})")
                else:
                    nokaspi_rows.append([folder, doc])

    out_xlsx = os.path.join(tempfile.gettempdir(),
                            f"Kaspi_Result_{datetime.now():%Y%m%d_%H%M}.xlsx")
    rows_written = build_excel(kaspi_rows, nokaspi_rows, total_docs, excluded, out_xlsx)

    return {
        "to_addr": sender_addr,
        "to_name": sender_name,
        "subject": subject,
        "msg_id": msg.get("Message-ID", ""),
        "xlsx": out_xlsx,
        "total_docs": total_docs,
        "excluded": excluded,
        "kaspi_n": len(kaspi_rows),
        "nokaspi_n": len(nokaspi_rows),
        "rows_written": rows_written,
        "flagged": flagged,
    }


def send_drive_error(res):
    """Tell the sender their Drive link wasn't accessible, with the fix."""
    body = (
        "Hello,\n\n"
        "We received your email but could not open the Google Drive file(s) you linked.\n\n"
        "Most likely the file isn't shared publicly. Please fix it like this:\n"
        "  1. In Google Drive, right-click the file → Share.\n"
        "  2. Under 'General access', choose 'Anyone with the link'.\n"
        "  3. Re-send the email with the link (subject must contain 'Kaspi').\n\n"
        "Technical detail:\n - " + "\n - ".join(res["errors"]) + "\n"
    )
    em = EmailMessage()
    em["From"] = GMAIL_USER
    em["To"] = res["to_addr"]
    em["Subject"] = "Kaspi — could not open your Drive file"
    em.set_content(body)
    ctx = ssl.create_default_context()
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=ctx) as smtp:
        smtp.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        smtp.send_message(em)


# ---- Send reply ------------------------------------------------------------
def send_reply(res):
    ok = (res["total_docs"] == res["rows_written"])
    body = (
        "Hello,\n\n"
        "Your Kaspi documents have been processed automatically.\n\n"
        f"Documents found (Excel files excluded): {res['total_docs']}\n"
        f"Excel files excluded: {res['excluded']}\n"
        f'Rows written ("Kaspi" tab): {res["kaspi_n"]}\n'
        f'Rows written ("No Kaspi" tab): {res["nokaspi_n"]}\n'
        f"Validation (documents = rows): {'PASSED' if ok else 'MISMATCH — please check'}\n"
    )
    if res["flagged"]:
        body += ("\nDocuments needing manual review (photo OCR is imperfect on\n"
                 "screen photos — please verify these against the originals):\n - "
                 + "\n - ".join(res["flagged"]) + "\n")
    body += "\nThe result file is attached.\n"

    em = EmailMessage()
    em["From"] = GMAIL_USER
    em["To"] = res["to_addr"]
    # Distinct subject so results don't pile into the incoming Gmail thread.
    em["Subject"] = "Kaspi Result — " + (res["subject"] or "receipts")
    em.set_content(body)

    with open(res["xlsx"], "rb") as f:
        em.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="Kaspi_Result.xlsx",
        )

    ctx = ssl.create_default_context()
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=ctx) as smtp:
        smtp.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        smtp.send_message(em)


# ---- Main ------------------------------------------------------------------
def main():
    ctx = ssl.create_default_context()
    imap = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT, ssl_context=ctx)
    imap.login(GMAIL_USER, GMAIL_APP_PASSWORD)
    imap.select("INBOX")

    # unseen emails whose subject contains the keyword.
    # Use UIDs (stable identifiers) so threading / deletions never cause mix-ups.
    typ, data = imap.uid('search', None, 'UNSEEN', 'SUBJECT', f'"{SUBJECT_KEYWORD}"')
    ids = data[0].split() if data and data[0] else []
    print(f"Found {len(ids)} candidate message(s).")

    processed = 0
    for uid in ids:
        typ, msg_data = imap.uid('fetch', uid, "(RFC822)")
        if not msg_data or msg_data[0] is None:
            continue
        raw = msg_data[0][1]
        try:
            res = process_message(raw)
        except Exception as e:
            print(f"  [error] message {uid!r}: {e}", file=sys.stderr)
            continue

        if res is None:
            print(f"  message {uid!r}: subject matched but no zip/Drive file — leaving unread.")
            continue

        if res.get("drive_error_only"):
            send_drive_error(res)
            imap.uid('store', uid, "+FLAGS", "\\Seen")   # notified sender; don't retry endlessly
            print(f"  message {uid!r}: Drive link not accessible — emailed sender the fix.")
            continue

        send_reply(res)
        imap.uid('store', uid, "+FLAGS", "\\Seen")   # mark done only after a successful reply
        processed += 1
        print(f"  Replied to {res['to_addr']}: {res['total_docs']} docs, "
              f"{res['rows_written']} rows, {len(res['flagged'])} flagged.")

    imap.logout()
    print(f"Done. Processed {processed} email(s).")


if __name__ == "__main__":
    main()
